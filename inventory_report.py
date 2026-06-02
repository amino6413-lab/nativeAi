import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

EXCEL_PATH = "물량표.xlsx"


def create_sample_excel():
    data = [
        (date(2026, 6, 1), "화장품A", 1200, "서울창고"),
        (date(2026, 6, 1), "화장품B",  850, "부산창고"),
        (date(2026, 6, 1), "화장품C",  630, "서울창고"),
        (date(2026, 6, 2), "화장품A",  970, "인천창고"),
        (date(2026, 6, 2), "화장품B", 1100, "서울창고"),
        (date(2026, 6, 2), "화장품C",  450, "부산창고"),
        (date(2026, 6, 3), "화장품A",  800, "부산창고"),
        (date(2026, 6, 3), "화장품B",  760, "인천창고"),
        (date(2026, 6, 3), "화장품C",  990, "서울창고"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "물량표"

    headers = ["날짜", "품목", "생산량", "창고"]
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=row[0]).number_format = "YYYY-MM-DD"
        ws.cell(row=row_idx, column=2, value=row[1])
        ws.cell(row=row_idx, column=3, value=row[2])
        ws.cell(row=row_idx, column=4, value=row[3])

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    wb.save(EXCEL_PATH)
    print(f"샘플 엑셀 생성 완료: {EXCEL_PATH}\n")


def print_report():
    df = pd.read_excel(EXCEL_PATH)
    df["날짜"] = pd.to_datetime(df["날짜"]).dt.strftime("%Y-%m-%d")

    col_widths = {col: max(len(col), df[col].astype(str).str.len().max()) for col in df.columns}

    def row_str(values):
        return " | ".join(str(v).ljust(col_widths[col]) for col, v in zip(df.columns, values))

    separator = "-+-".join("-" * col_widths[col] for col in df.columns)

    print("=" * len(separator))
    print("           물량표 리포트")
    print("=" * len(separator))
    print(row_str(df.columns))
    print(separator)
    for _, row in df.iterrows():
        print(row_str(row))
    print(separator)

    # 품목별 소계
    subtotals = df.groupby("품목")["생산량"].sum()
    print("\n[품목별 생산량 합계]")
    for item, total in subtotals.items():
        print(f"  {item}: {total:,} 개")

    grand_total = df["생산량"].sum()
    print(f"\n  총 생산량 합계: {grand_total:,} 개")
    print("=" * len(separator))


if __name__ == "__main__":
    create_sample_excel()
    print_report()